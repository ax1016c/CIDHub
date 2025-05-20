import os
import shutil
import tkinter as tk
from tkinter import filedialog, ttk, scrolledtext, messagebox
from pathlib import Path
import openpyxl
import threading
from queue import Queue, Empty

class FileFinder:
    def __init__(self, root):
        self.root = root
        self.root.title("Buscador y Organizador de Archivos PDF")
        self.root.geometry("800x680")
        self.root.resizable(True, True)

        # Variables para rutas y configuraciones
        self.excel_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.id_column = tk.StringVar(value="A")
        self.name_column = tk.StringVar(value="B")
        self.source_dir = tk.StringVar()

        # Control de threading y estado
        self.stop_requested = False
        self.processing_thread = None
        self.log_queue = Queue()

        # Widgets
        self._create_widgets()
        self.root.after(100, self._poll_log_queue)

    def _create_widgets(self):
        frame = ttk.Frame(self.root, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        # Selección de Excel
        excel_frame = ttk.LabelFrame(frame, text="Archivo Excel (.xlsx)", padding=5)
        excel_frame.pack(fill=tk.X, pady=5)
        ttk.Button(excel_frame, text="Seleccionar Excel...", command=self._select_excel).pack(side=tk.LEFT)
        ttk.Label(excel_frame, textvariable=self.excel_path, wraplength=600).pack(side=tk.LEFT, padx=5)

        # Hoja y columnas
        config_frame = ttk.LabelFrame(frame, text="Hoja y Columnas", padding=5)
        config_frame.pack(fill=tk.X, pady=5)
        ttk.Label(config_frame, text="Hoja:").grid(row=0, column=0, sticky="w")
        self.sheet_cb = ttk.Combobox(config_frame, textvariable=self.sheet_name, state="readonly", width=30)
        self.sheet_cb.grid(row=0, column=1, padx=5, pady=2)
        ttk.Label(config_frame, text="Columna ID:").grid(row=1, column=0, sticky="w")
        ttk.Entry(config_frame, textvariable=self.id_column, width=5).grid(row=1, column=1, sticky="w", padx=5)
        ttk.Label(config_frame, text="Columna NOMBRE:").grid(row=2, column=0, sticky="w")
        ttk.Entry(config_frame, textvariable=self.name_column, width=5).grid(row=2, column=1, sticky="w", padx=5)

        # Selección de carpeta origen
        dir_frame = ttk.LabelFrame(frame, text="Directorio Fuente", padding=5)
        dir_frame.pack(fill=tk.X, pady=5)
        ttk.Button(dir_frame, text="Seleccionar Carpeta...", command=self._select_source_dir).pack(side=tk.LEFT)
        ttk.Label(dir_frame, textvariable=self.source_dir, wraplength=600).pack(side=tk.LEFT, padx=5)

        # Botones de acción
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=10)
        self.btn_process = ttk.Button(btn_frame, text="Procesar PDF", command=self._start_thread)
        self.btn_process.pack(side=tk.LEFT, padx=5)
        self.btn_stop = ttk.Button(btn_frame, text="Detener", command=self._request_stop, state=tk.DISABLED)
        self.btn_stop.pack(side=tk.LEFT)

        # Área de logs
        log_frame = ttk.LabelFrame(frame, text="Registro", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.txt_log = scrolledtext.ScrolledText(log_frame, height=15, state=tk.DISABLED)
        self.txt_log.pack(fill=tk.BOTH, expand=True)

    def _poll_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                self._append_log(msg)
        except Empty:
            pass
        finally:
            self.root.after(100, self._poll_log_queue)

    def _append_log(self, msg):
        self.txt_log.config(state=tk.NORMAL)
        self.txt_log.insert(tk.END, msg + "\n")
        self.txt_log.see(tk.END)
        self.txt_log.config(state=tk.DISABLED)

    def _log(self, msg):
        self.log_queue.put(msg)

    def _select_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.excel_path.set(path)
            self._load_sheets(path)

    def _load_sheets(self, path):
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            names = wb.sheetnames
            wb.close()
            self.sheet_cb['values'] = names
            if names:
                self.sheet_name.set(names[0])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar las hojas: {e}")

    def _select_source_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.source_dir.set(path)

    def _start_thread(self):
        if not (self.excel_path.get() and self.sheet_name.get() and self.id_column.get() and self.name_column.get() and self.source_dir.get()):
            messagebox.showwarning("Datos incompletos", "Seleccione Excel, hoja, columnas ID y NOMBRE, y carpeta origen.")
            return
        self.stop_requested = False
        self.btn_process.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)
        self.txt_log.config(state=tk.NORMAL)
        self.txt_log.delete(1.0, tk.END)
        self.txt_log.config(state=tk.DISABLED)
        self.processing_thread = threading.Thread(target=self._process, daemon=True)
        self.processing_thread.start()

    def _request_stop(self):
        if messagebox.askyesno("Detener", "¿Desea detener el proceso? Los archivos copiados permanecerán."):
            self.stop_requested = True
            self._log("Detención solicitada...")

    def _create_dest_folder(self):
        base = Path(self.source_dir.get()) / Path(self.excel_path.get()).stem
        counter = 1
        dest = base
        while dest.exists():
            dest = Path(f"{base}_{counter}")
            counter += 1
        dest.mkdir(parents=True)
        return dest

    def _read_excel(self):
        self._log(f"Leyendo Excel: {self.excel_path.get()}, Hoja: {self.sheet_name.get()}")
        wb = openpyxl.load_workbook(self.excel_path.get(), read_only=True)
        sheet = wb[self.sheet_name.get()]
        col_id = openpyxl.utils.column_index_from_string(self.id_column.get().upper())
        col_name = openpyxl.utils.column_index_from_string(self.name_column.get().upper())
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            id_val = row[col_id-1]
            name_val = row[col_name-1]
            if id_val and name_val:
                data.append((str(id_val).strip(), str(name_val).strip()))
        wb.close()
        self._log(f"Registros leídos: {len(data)}")
        return data

    def _process(self):
        try:
            dest_root = self._create_dest_folder()
            records = self._read_excel()
            total_copied = 0

            for idx, (id_val, name_val) in enumerate(records, 1):
                if self.stop_requested:
                    break
                self._log(f"[{idx}/{len(records)}] ID={id_val}, NOMBRE={name_val}")
                folder = dest_root / name_val
                folder.mkdir(exist_ok=True)
                # Buscar PDFs
                found = []
                for root, _, files in os.walk(self.source_dir.get()):
                    if self.stop_requested:
                        break
                    for f in files:
                        if not f.lower().endswith('.pdf'):
                            continue
                        if id_val.lower() in Path(f).stem.lower():
                            found.append(Path(root) / f)
                    
                if not found:
                    self._log(f"  ✗ No se encontró PDF para ID {id_val}")
                else:
                    for file_path in found:
                        if self.stop_requested:
                            break
                        try:
                            shutil.copy2(file_path, folder / file_path.name)
                            self._log(f"  ✓ Copiado {file_path.name} a {folder}")
                            total_copied += 1
                        except Exception as e:
                            self._log(f"  ✗ Error copiando {file_path.name}: {e}")
            # Resumen
            self._log("\n--- Resumen ---")
            self._log(f"Total registros: {len(records)}")
            self._log(f"Total PDFs copiados: {total_copied}")
            if self.stop_requested:
                self._log("Proceso detenido por el usuario.")
            else:
                self._log("Proceso completado.")
        except Exception as e:
            self._log(f"Error en procesamiento: {e}")
        finally:
            self.root.after(0, self._reset_ui)

    def _reset_ui(self):
        self.btn_process.config(state=tk.NORMAL)
        self.btn_stop.config(state=tk.DISABLED)

if __name__ == "__main__":
    root = tk.Tk()
    app = FileFinder(root)
    root.mainloop()
