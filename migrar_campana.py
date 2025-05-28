#!/usr/bin/env python3
"""
Script para migrar datos de campaña.xlsx al formato de Plantilla.xlsx
Llena tanto la hoja de Clientes como la de Leads con la información adaptada
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
from datetime import datetime

def clean_phone_number(phone):
    """Limpia y formatea números telefónicos"""
    if pd.isna(phone):
        return ""
    phone_str = str(int(phone)) if isinstance(phone, float) else str(phone)
    # Eliminar caracteres no numéricos excepto +
    phone_clean = re.sub(r'[^\d+]', '', phone_str)
    return phone_clean

def clean_vendor_name(vendor):
    """Limpia nombres de vendedores eliminando sufijos como .NUEVO"""
    if pd.isna(vendor):
        return ""
    # Eliminar sufijos como ".NUEVO", ".   NUEVO", etc.
    cleaned = re.sub(r'\.\s*(NUEVO|nuevo).*$', '', str(vendor))
    return cleaned.strip()

def normalize_stage(stage):
    """Normaliza las etapas de ventas"""
    if pd.isna(stage):
        return ""
    
    stage_mapping = {
        "Nuevo": "Nuevo",
        "Reconocimiento de necesidades": "Reconocimiento de necesidades",
        "Reconocimiento de necesidades ": "Reconocimiento de necesidades",  # Con espacio extra
        "Presentación de cotización": "Presentación de cotización",
        "Negociación final": "Negociación final",
        "Negociación Final": "Negociación final",  # Normalizar capitalización
        "Ganado": "Ganado",
        "Perdido": "Perdido"
    }
    
    return stage_mapping.get(str(stage).strip(), str(stage).strip())

# MODIFICADO: Esta función ya no es necesaria para determinar el tipo de cliente en la salida,
# ya que se ha fijado a persona física (2). Se mantiene por si se requiere en el futuro
# para otros propósitos o para entender la lógica original.
# def determine_client_type(title, contact_name):
#     """Determina si es persona física (2) o moral (1) basado en título y nombre"""
#     if pd.isna(title) and pd.isna(contact_name):
#         return 2  # Por defecto persona física
    
#     title_str = str(title).lower() if not pd.isna(title) else ""
#     contact_str = str(contact_name).lower() if not pd.isna(contact_name) else ""
    
#     company_indicators = ['s.a.', 'sa', 'srl', 'sc', 'asociación', 'fundación', 
#                          'instituto', 'centro', 'hospital', 'clínica', 'consultorio',
#                          'grupo', 'corporativo', 'empresa']
    
#     full_text = f"{title_str} {contact_str}"
    
#     for indicator in company_indicators:
#         if indicator in full_text:
#             return 1  # Persona moral
    
#     return 2  # Persona física

def create_display_name(title, contact_name):
    """Crea el nombre mostrado combinando título y nombre de contacto"""
    if pd.isna(contact_name):
        return ""
    
    if pd.isna(title):
        return str(contact_name).strip()
    
    return f"{str(title).strip()} {str(contact_name).strip()}".strip()

def generate_opportunity_description(contact_name, stage, vendor): # stage and vendor are no longer used by the new logic
    """
    MODIFICADO: Genera una descripción de oportunidad basada solo en el nombre de contacto.
    Elimina 'Oportunidad - ' y solo deja los nombres.
    """
    if pd.isna(contact_name):
        return ""  # Si no hay nombre de contacto, devuelve una cadena vacía
    return str(contact_name).strip()

# MODIFICADO: Esta función ya no es necesaria para determinar las etiquetas,
# ya que se ha fijado a "MRC3D 1". Se mantiene por si se requiere en el futuro
# para otros propósitos o para entender la lógica original.
# def determine_tags_from_stage(stage):
#     """Determina etiquetas basadas en la etapa de ventas"""
#     if pd.isna(stage):
#         return "Prospecto"
    
#     stage_str = str(stage).lower()
    
#     if "nuevo" in stage_str:
#         return "Prospecto, Nuevo"
#     elif "reconocimiento" in stage_str:
#         return "Demo, Análisis de necesidades"
#     elif "cotización" in stage_str or "presentación" in stage_str:
#         return "Cotización, Propuesta"
#     elif "negociación" in stage_str:
#         return "Cotización, Negociación"
#     elif "ganado" in stage_str:
#         return "Cliente, Ganado"
#     elif "perdido" in stage_str:
#         return "Perdido"
#     else:
#         return "Prospecto"

def migrate_campaign_to_template(campaign_file, template_file, output_file):
    """
    Función principal que migra los datos de campaña al formato de plantilla
    """
    
    print("Leyendo archivo de campaña...")
    campaign_df = pd.read_excel(campaign_file, sheet_name='ALL')
    
    print(f"Se encontraron {len(campaign_df)} registros en la campaña")
    
    print("Limpiando y preparando datos...")
    campaign_df['Teléfono_Clean'] = campaign_df['Teléfono'].apply(clean_phone_number)
    campaign_df['Vendedor_Clean'] = campaign_df['Vendedor'].apply(clean_vendor_name)
    campaign_df['Etapa_Clean'] = campaign_df['Etapa'].apply(normalize_stage)
    campaign_df['Nombre_Mostrado'] = campaign_df.apply(
        lambda row: create_display_name(row['Título'], row['Contacto']), axis=1
    )
    # MODIFICADO: Ya no se pre-calcula 'Tipo_Cliente' aquí, se asigna directamente.
    # campaign_df['Tipo_Cliente'] = campaign_df.apply(
    #     lambda row: determine_client_type(row['Título'], row['Contacto']), axis=1
    # )
    
    print("Cargando plantilla...")
    template_wb = load_workbook(template_file)
    
    print("Preparando datos para hoja Clientes...")
    clientes_data = []
    
    for _, row in campaign_df.iterrows():
        cliente_row = [
            2,  # MODIFICADO: Es una empresa (1 Moral y 2 física) -> Todos como persona física (2)
            row['Nombre_Mostrado'],  # Nombre mostrado /razon social
            row['Teléfono_Clean'],  # Telefono Celular
            "",  # email (vacío)
            row['Contacto'] if not pd.isna(row['Contacto']) else "",  # Nombre del contacto
            "Teléfono",  # Medio
            "",  # Calle
            "",  # Calle2 / referencias / Colonia
            "",  # Casa/Num Ext
            "",  # Puerta/Num Interior
            "",  # C.P.
            "",  # Ciudad
            "",  # Estado/Nombre mostrado
            "México",  # country_id/Pais
            row['Vendedor_Clean'],  # Vendedor
            1,  # Rango de cliente (valor secuencial)
            "",  # RFC
            ""   # Regimen fiscal
        ]
        clientes_data.append(cliente_row)
    
    print("Preparando datos para hoja Leads...")
    leads_data = []
    
    for _, row in campaign_df.iterrows():
        # MODIFICADO: Llamada a la función generate_opportunity_description actualizada
        oportunidad_desc = generate_opportunity_description(
            row['Contacto'], row['Etapa_Clean'], row['Vendedor_Clean']
        )
        # MODIFICADO: Etiquetas fijas
        etiquetas = "MRC3D 1"
        
        lead_row = [
            oportunidad_desc,  # Oportunidad Descripción (MODIFICADO: solo nombres)
            row['Vendedor_Clean'],  # Vendedor
            row['Contacto'] if not pd.isna(row['Contacto']) else "",  # Nombre del contacto
            row['Etapa_Clean'],  # Etapa
            "",  # Ingreso esperado
            "",  # Tipo de compra
            "",  # Producto de interés/Nombre
            etiquetas,  # Etiquetas, seleccionar las que correspondan (MODIFICADO: "MRC3D 1")
            "Facebook",  # Origen/Nombre de la fuente (MODIFICADO: "Facebook")
            "",  # Recomendado por
            "Ventas"  # Equipo de ventas/Nombre en pantalla
        ]
        leads_data.append(lead_row)
    
    print("Escribiendo datos en hoja Clientes...")
    clientes_ws = template_wb['Clientes']
    
    if clientes_ws.max_row > 1: # Check if there's data beyond headers
        clientes_ws.delete_rows(2, clientes_ws.max_row -1) # Delete existing data rows

    for i, cliente in enumerate(clientes_data, start=2): # Start from row 2 (after headers)
        for j, valor in enumerate(cliente, start=1):
            clientes_ws.cell(row=i, column=j, value=valor)
    
    print("Escribiendo datos en hoja Leads...")
    leads_ws = template_wb['Leads']

    if leads_ws.max_row > 1: # Check if there's data beyond headers
        leads_ws.delete_rows(2, leads_ws.max_row -1) # Delete existing data rows
    
    for i, lead in enumerate(leads_data, start=2): # Start from row 2 (after headers)
        for j, valor in enumerate(lead, start=1):
            leads_ws.cell(row=i, column=j, value=valor)
    
    print(f"Guardando archivo resultado: {output_file}")
    template_wb.save(output_file)
    
    print("¡Migración completada exitosamente!")
    print(f"- {len(clientes_data)} registros añadidos a la hoja Clientes")
    print(f"- {len(leads_data)} registros añadidos a la hoja Leads")
    
    print("\n=== ESTADÍSTICAS DE MIGRACIÓN ===")
    print(f"Vendedores únicos: {campaign_df['Vendedor_Clean'].nunique()}")
    print(f"Etapas de venta encontradas (original): {campaign_df['Etapa'].nunique()}") # Original stages
    print("Distribución por etapa (normalizada):")
    for etapa, count in campaign_df['Etapa_Clean'].value_counts().items():
        print(f"  - {etapa}: {count} registros")

def main():
    """Función principal del script"""
    campaign_file = "campaña.xlsx"
    template_file = "Plantilla.xlsx"
    output_file = f"Plantilla_Migrada_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    try:
        migrate_campaign_to_template(campaign_file, template_file, output_file)
        print(f"\n✅ Proceso completado. Archivo generado: {output_file}")
        
    except FileNotFoundError as e:
        print(f"❌ Error: No se pudo encontrar el archivo {e.filename}")
        print("Asegúrate de que los archivos 'campaña.xlsx' y 'Plantilla.xlsx' estén en el mismo directorio que este script.")
        
    except Exception as e:
        print(f"❌ Error inesperado: {str(e)}")
        print("Por favor, verifica que los archivos tengan el formato correcto y que las hojas 'Clientes' y 'Leads' existan en la plantilla.")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

# ==================== SCRIPT SIMPLE ALTERNATIVO (NO MODIFICADO CON LOS NUEVOS REQUISITOS) ====================
"""
Si prefieres un enfoque más directo, puedes usar esta versión simplificada:

import pandas as pd
from openpyxl import load_workbook

# Leer campaña
df = pd.read_excel('campaña.xlsx', sheet_name='ALL')

# Cargar plantilla
wb = load_workbook('Plantilla.xlsx')

# Llenar Clientes
clientes_ws = wb['Clientes']
for i, (_, row) in enumerate(df.iterrows(), start=2):
    clientes_ws[f'A{i}'] = 2  # Persona física
    clientes_ws[f'B{i}'] = f"{row.get('Título', '')} {row.get('Contacto', '')}".strip()
    clientes_ws[f'C{i}'] = str(row.get('Teléfono', ''))
    clientes_ws[f'E{i}'] = row.get('Contacto', '')
    clientes_ws[f'O{i}'] = row.get('Vendedor', '')

# Llenar Leads  
leads_ws = wb['Leads']
for i, (_, row) in enumerate(df.iterrows(), start=2):
    leads_ws[f'A{i}'] = f"Oportunidad - {row.get('Contacto', '')}"
    leads_ws[f'B{i}'] = row.get('Vendedor', '')
    leads_ws[f'C{i}'] = row.get('Contacto', '')
    leads_ws[f'D{i}'] = row.get('Etapa', '')

# Guardar
wb.save('Resultado_Simple.xlsx')
"""