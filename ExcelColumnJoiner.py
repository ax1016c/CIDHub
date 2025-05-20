import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
from pathlib import Path

class ExcelColumnJoiner(tk.Tk):
    def __init__(self):
        super().__init__()
        
        # Configuración de la ventana principal
        self.title("Unificador y Ordenador de Columnas Excel")
        self.geometry("600x450") # Reducimos altura ya que eliminamos controles innecesarios
        self.resizable(True, True)
        
        # Variables para almacenar información
        self.file_path = tk.StringVar()
        self.source_column = tk.StringVar()
        self.target_column = tk.StringVar()
        self.status_message = tk.StringVar()
        self.selected_sheet = tk.StringVar()
        self.sheet_names = []
        
        # Crear el marco principal
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # --- Sección de selección de archivo ---
        file_frame = ttk.LabelFrame(main_frame, text="1. Selección de Archivo", padding="10")
        file_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(file_frame, text="Archivo Excel:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(file_frame, textvariable=self.file_path, width=50, state="readonly").grid(row=0, column=1, sticky=tk.W+tk.E, padx=5, pady=5)
        ttk.Button(file_frame, text="Examinar...", command=self.browse_file).grid(row=0, column=2, padx=5, pady=5)
        
        # --- Sección de selección de hoja ---
        sheet_frame = ttk.LabelFrame(main_frame, text="2. Selección de Hoja", padding="10")
        sheet_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(sheet_frame, text="Hoja de trabajo:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.sheet_combobox = ttk.Combobox(sheet_frame, textvariable=self.selected_sheet, state="readonly", width=47)
        self.sheet_combobox.grid(row=0, column=1, sticky=tk.W+tk.E, padx=5, pady=5)
        
        # --- Sección de configuración de columnas para UNIÓN ---
        columns_frame = ttk.LabelFrame(main_frame, text="3. Configuración de Unión y Ordenamiento de Columnas", padding="10")
        columns_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(columns_frame, text="Columna Fuente (ej: K):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(columns_frame, textvariable=self.source_column, width=10).grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        
        ttk.Label(columns_frame, text="Columna Destino (ej: L):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(columns_frame, textvariable=self.target_column, width=10).grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        
        # Botón para el proceso completo
        ttk.Button(columns_frame, text="Unir y Ordenar Columnas", command=self.join_and_sort_columns).grid(row=0, column=2, rowspan=2, padx=10, pady=5, sticky="ns")

        # --- Sección de estado ---
        status_frame = ttk.LabelFrame(main_frame, text="Estado", padding="10")
        status_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.status_label = ttk.Label(status_frame, textvariable=self.status_message, wraplength=550, justify=tk.LEFT)
        self.status_label.pack(fill=tk.X)
        
        # Mensajes de guía inicial
        self.status_message.set("Por favor seleccione un archivo Excel (.xlsx) para comenzar.")

    def browse_file(self):
        """Abre un diálogo para seleccionar un archivo Excel."""
        file_path_val = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        
        if file_path_val:
            self.file_path.set(file_path_val)
            self.status_message.set(f"Archivo seleccionado: {os.path.basename(file_path_val)}")
            self.load_sheet_names()

    def load_sheet_names(self):
        """Carga los nombres de las hojas del archivo Excel seleccionado."""
        try:
            if not self.file_path.get():
                return
            workbook = openpyxl.load_workbook(self.file_path.get(), read_only=True)
            self.sheet_names = workbook.sheetnames
            self.sheet_combobox['values'] = self.sheet_names
            
            if self.sheet_names:
                self.selected_sheet.set(self.sheet_names[0])
            else:
                self.selected_sheet.set("")
                
            workbook.close()
        except Exception as e:
            self.status_message.set(f"Error al cargar las hojas: {str(e)}")
            messagebox.showerror("Error", f"No se pudieron cargar las hojas del archivo:\n{str(e)}")
            self.sheet_combobox['values'] = []
            self.selected_sheet.set("")

    def ajustar_ancho_columnas_hoja(self, worksheet):
        """Ajusta el ancho de las columnas en la hoja dada al máximo contenido."""
        anchos_maximos_columna = {} 
        for fila in worksheet.iter_rows(): 
            for celda in fila:
                if celda.value is not None: 
                    letra_columna = celda.column_letter
                    longitud_texto = len(str(celda.value))
                    if longitud_texto > anchos_maximos_columna.get(letra_columna, 0):
                        anchos_maximos_columna[letra_columna] = longitud_texto
        
        for letra_columna, longitud_maxima in anchos_maximos_columna.items():
            ancho_ajustado = longitud_maxima + 2 
            worksheet.column_dimensions[letra_columna].width = ancho_ajustado

    def join_and_sort_columns(self):
        """Ejecuta el proceso completo: unir columnas y crear directamente el archivo ordenado con solo las columnas deseadas."""
        if not self.file_path.get():
            self.status_message.set("Error: No se ha seleccionado ningún archivo.")
            messagebox.showerror("Error", "Por favor seleccione un archivo Excel primero.")
            return
        
        if not self.selected_sheet.get():
            self.status_message.set("Error: No se ha seleccionado ninguna hoja de trabajo.")
            messagebox.showerror("Error", "Por favor seleccione una hoja de trabajo.")
            return
            
        source_col = self.source_column.get().strip().upper()
        target_col = self.target_column.get().strip().upper()
        
        if not source_col or not target_col:
            self.status_message.set("Error: Debe especificar ambas columnas (fuente y destino).")
            messagebox.showerror("Error", "Por favor ingrese letras válidas para las columnas fuente y destino.")
            return
            
        if not (source_col.isalpha() and target_col.isalpha()):
            self.status_message.set("Error: Las columnas deben ser letras válidas (A-Z).")
            messagebox.showerror("Error", "Las columnas deben especificarse como letras (A-Z).")
            return
            
        try:
            workbook = openpyxl.load_workbook(self.file_path.get())
            sheet_name = self.selected_sheet.get()
            worksheet = workbook[sheet_name]
            
            # Realizar la unión de columnas
            rows_processed = 0
            rows_modified = 0
            
            for row_idx in range(1, worksheet.max_row + 1):
                source_cell_val = worksheet[f"{source_col}{row_idx}"].value
                target_cell_obj = worksheet[f"{target_col}{row_idx}"]
                
                rows_processed += 1
                
                if source_cell_val is not None and str(source_cell_val).strip() != "":
                    if target_cell_obj.value is not None:
                        target_cell_obj.value = str(source_cell_val) + str(target_cell_obj.value)
                    else:
                        target_cell_obj.value = source_cell_val
                    rows_modified += 1
            
            # Extraer datos de la hoja después de la unión para ordenar columnas
            data = list(worksheet.iter_rows())  # Usar iter_rows para obtener celdas

            if not data:
                self.status_message.set(f"Hoja '{sheet_name}' está vacía. Nada que ordenar.")
                workbook.close()
                return

            header_cells = data[0]
            original_headers = [cell.value for cell in header_cells]
            
            # Extraer todas las filas de datos (valores)
            data_rows_values = []
            for row_cells in data[1:]:
                data_rows_values.append([cell.value for cell in row_cells])

            original_header_to_index = {header: i for i, header in enumerate(original_headers)}
            
            column_data_by_original_index = []
            for i in range(len(original_headers)):
                col_values = [original_headers[i]] # Header
                for row_data_tuple in data_rows_values:
                    if i < len(row_data_tuple):
                        col_values.append(row_data_tuple[i])
                    else:
                        col_values.append(None) 
                column_data_by_original_index.append(col_values)

            # Orden deseado para las columnas - solo incluir estas columnas
            desired_order_spec = [
                'Folio', 'UUID', 'Fecha Emision', 'FormaDePago', 'Condicion de Pago', 
                'SubTotal', 'Moneda', 'Total', 'Tipo', 'LugarDeExpedicion', 'Nombre Emisor'
            ]

            new_ordered_columns_data = []
            final_header_order_for_output = []

            # Solo incluir columnas en el orden especificado
            for desired_header in desired_order_spec:
                original_col_idx_to_use = -1
                actual_header_found = None

                # Caso especial para 'Folio' que puede venir de 'SerieFolio'
                if desired_header == 'Folio':
                    if 'Folio' in original_header_to_index:
                        original_col_idx_to_use = original_header_to_index['Folio']
                        actual_header_found = 'Folio'
                    elif 'SerieFolio' in original_header_to_index:
                        original_col_idx_to_use = original_header_to_index['SerieFolio']
                        actual_header_found = 'SerieFolio'  # El que se encontró
                # Búsqueda normal
                elif desired_header.lower() in [h.lower() if h else '' for h in original_headers]:
                    # Buscar sin importar mayúsculas/minúsculas
                    for idx, header in enumerate(original_headers):
                        if header and header.lower() == desired_header.lower():
                            original_col_idx_to_use = idx
                            actual_header_found = header
                            break
                
                if original_col_idx_to_use != -1:
                    # Copiar la columna de datos original
                    col_data_copy = list(column_data_by_original_index[original_col_idx_to_use])
                    # Establecer el encabezado en la copia al deseado ('Folio' si vino de 'SerieFolio')
                    col_data_copy[0] = desired_header 
                    new_ordered_columns_data.append(col_data_copy)
                    final_header_order_for_output.append(desired_header)

            # Crear un nuevo libro y una nueva hoja para el resultado ordenado
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = sheet_name

            if new_ordered_columns_data:
                # Determinar el número de filas a escribir (longitud de la columna más larga)
                max_rows_to_write = 0
                if new_ordered_columns_data:
                    max_rows_to_write = max(len(col) for col in new_ordered_columns_data)

                num_cols_to_write = len(new_ordered_columns_data)

                # Transponer y escribir los datos
                for r_idx in range(max_rows_to_write):
                    row_to_write = []
                    for c_idx in range(num_cols_to_write):
                        if r_idx < len(new_ordered_columns_data[c_idx]):
                            row_to_write.append(new_ordered_columns_data[c_idx][r_idx])
                        else:
                            row_to_write.append(None)  # Rellenar si la columna es más corta
                    new_sheet.append(row_to_write)
            
            # Ajustar el ancho de las columnas en la hoja final
            self.ajustar_ancho_columnas_hoja(new_sheet)

            # Crear un nombre para el archivo ordenado
            original_path = Path(self.file_path.get())
            final_filename = original_path.stem + "_ordenado" + original_path.suffix
            final_file_path = original_path.parent / final_filename
            
            # Guardar el archivo ordenado
            new_workbook.save(final_file_path)
            new_workbook.close()
            workbook.close()
            
            # Mensaje de éxito final
            final_success_message = (
                f"Proceso completado exitosamente.\n"
                f"- Columnas unidas y solo se incluyeron las columnas especificadas en el orden requerido.\n"
                f"- Archivo final guardado como: {final_filename}\n"
                f"- Los anchos de columna han sido ajustados.\n"
            )
            self.status_message.set(final_success_message)
            messagebox.showinfo("Proceso Completado", final_success_message)
            
        except Exception as e:
            error_msg = f"Error durante el proceso: {str(e)}"
            self.status_message.set(error_msg)
            messagebox.showerror("Error en Proceso", error_msg)
            if 'workbook' in locals() and workbook:
                try: workbook.close()
                except: pass
            if 'new_workbook' in locals() and new_workbook:
                try: new_workbook.close()
                except: pass

if __name__ == "__main__":
    app = ExcelColumnJoiner()
    app.mainloop()